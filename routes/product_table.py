import os
import json
from datetime import datetime
from flask import Blueprint, request, render_template, jsonify, send_file, current_app
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import glob
from pathlib import Path

product_table_bp = Blueprint('product_table', __name__, url_prefix='/product-table')


@product_table_bp.route('/', methods=['GET', 'POST'])
def product_table():
    """亚马逊商品表格生成页面"""
    if request.method == 'GET':
        return render_template('product_table.html')
    
    # POST 处理 - 生成亚马逊商品表格
    try:
        return handle_amazon_table_generation()
    
    except Exception as e:
        return render_template('product_table.html', 
                             error=f"生成表格时出错: {str(e)}")

def handle_amazon_table_generation():
    """处理亚马逊模式的表格生成"""
    try:
        # 获取表单数据
        table_title = request.form.get('amazon_table_title', '亚马逊商品表格')
        batch_folder = request.form.get('amazon_batch_folder')
        export_format = request.form.get('amazon_export_format', 'excel')
        
        if not batch_folder:
            return render_template('product_table.html', 
                                 error="请选择批次文件夹")
        
        # 检查是否上传了模板文件
        if 'amazon_template' not in request.files:
            return render_template('product_table.html', 
                                 error="请上传亚马逊商品模板")
        
        template_file = request.files['amazon_template']
        if template_file.filename == '':
            return render_template('product_table.html', 
                                 error="请选择模板文件")
        
        # 解析模板文件
        variants = parse_template_file(template_file)
        
        # 获取批次文件夹中的商品
        products = extract_products_from_batch(batch_folder)
        
        # 匹配变体与商品
        matches = match_variants_with_products(variants, products)
        
        # 生成亚马逊格式的商品数据
        amazon_products = generate_amazon_product_data(matches, products)
        
        # 生成表格
        if export_format == 'excel':
            file_buffer = generate_amazon_excel_table(table_title, amazon_products)
            filename = f"{table_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                file_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        elif export_format == 'csv':
            file_buffer = generate_amazon_csv_table(table_title, amazon_products)
            filename = f"{table_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return send_file(
                file_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='text/csv'
            )
        
        else:
            return render_template('product_table.html', 
                                 error="不支持的导出格式")
    
    except Exception as e:
        return render_template('product_table.html', 
                             error=f"生成亚马逊表格时出错: {str(e)}")

def generate_excel_table(title, products):
    """生成Excel格式的商品表格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"
    
    # 设置样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='微软雅黑', size=11)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 表头
    headers = ['序号', '商品名称', '商品编码', '价格(元)', '库存数量', '商品描述']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 设置列宽
    column_widths = [8, 20, 15, 12, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 数据行
    for row, product in enumerate(products, 4):
        # 序号
        cell = ws.cell(row=row, column=1, value=row - 3)
        cell.font = content_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 商品名称
        cell = ws.cell(row=row, column=2, value=product.get('name', ''))
        cell.font = content_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        
        # 商品编码
        cell = ws.cell(row=row, column=3, value=product.get('code', ''))
        cell.font = content_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 价格
        price = product.get('price', 0)
        try:
            price = float(price)
        except (ValueError, TypeError):
            price = 0
        cell = ws.cell(row=row, column=4, value=price)
        cell.font = content_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # 库存数量
        stock = product.get('stock', 0)
        try:
            stock = int(stock)
        except (ValueError, TypeError):
            stock = 0
        cell = ws.cell(row=row, column=5, value=stock)
        cell.font = content_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 商品描述
        cell = ws.cell(row=row, column=6, value=product.get('description', ''))
        cell.font = content_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
    
    # 保存到内存
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


def generate_csv_table(title, products):
    """生成CSV格式的商品表格"""
    # 准备数据
    data = []
    for i, product in enumerate(products, 1):
        price = product.get('price', 0)
        try:
            price = float(price)
        except (ValueError, TypeError):
            price = 0
            
        stock = product.get('stock', 0)
        try:
            stock = int(stock)
        except (ValueError, TypeError):
            stock = 0
            
        data.append({
            '序号': i,
            '商品名称': product.get('name', ''),
            '商品编码': product.get('code', ''),
            '价格(元)': price,
            '库存数量': stock,
            '商品描述': product.get('description', '')
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存到内存
    file_buffer = BytesIO()
    df.to_csv(file_buffer, index=False, encoding='utf-8-sig')
    file_buffer.seek(0)
    
    return file_buffer


def generate_amazon_product_data(matches, products):
    """根据匹配结果生成亚马逊格式的商品数据"""
    amazon_products = []
    
    # 创建商品文件夹到商品信息的映射
    product_map = {product['folder_name']: product for product in products}
    
    for match in matches:
        if match['product_folder'] and match['product_folder'] in product_map:
            product = product_map[match['product_folder']]
            
            # 解析变体信息
            variant_parts = match['variant_info'].split(' - ')
            parent_sku = variant_parts[0] if len(variant_parts) > 0 else ''
            color = variant_parts[1] if len(variant_parts) > 1 else ''
            size = variant_parts[2] if len(variant_parts) > 2 else ''
            
            amazon_product = {
                'parent_sku': parent_sku,
                'child_sku': f"{parent_sku}_{color}_{size}".replace(' ', '_'),
                'product_name': product.get('name', ''),
                'color': color,
                'size': size,
                'folder_name': product.get('folder_name', ''),
                'image_count': product.get('image_count', 0),
                'preview_image': product.get('preview_image', ''),
                'batch_name': product.get('batch_name', ''),
                'created_time': product.get('created_time', '')
            }
            
            amazon_products.append(amazon_product)
    
    return amazon_products


def generate_amazon_excel_table(title, amazon_products):
    """生成亚马逊格式的Excel表格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "亚马逊商品信息"
    
    # 设置样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='微软雅黑', size=11)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ['序号', '父SKU', '子SKU', '商品名称', '颜色', '尺寸', '文件夹名称', '图片数量', '批次名称', '创建时间']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行
    for row, product in enumerate(amazon_products, 4):
        data = [
            row - 3,  # 序号
            product.get('parent_sku', ''),
            product.get('child_sku', ''),
            product.get('product_name', ''),
            product.get('color', ''),
            product.get('size', ''),
            product.get('folder_name', ''),
            product.get('image_count', 0),
            product.get('batch_name', ''),
            product.get('created_time', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = content_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    column_widths = [8, 20, 25, 25, 15, 15, 20, 12, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 保存到内存
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


def generate_amazon_csv_table(title, amazon_products):
    """生成亚马逊格式的CSV表格"""
    data = []
    for i, product in enumerate(amazon_products, 1):
        data.append({
            '序号': i,
            '父SKU': product.get('parent_sku', ''),
            '子SKU': product.get('child_sku', ''),
            '商品名称': product.get('product_name', ''),
            '颜色': product.get('color', ''),
            '尺寸': product.get('size', ''),
            '文件夹名称': product.get('folder_name', ''),
            '图片数量': product.get('image_count', 0),
            '批次名称': product.get('batch_name', ''),
            '创建时间': product.get('created_time', '')
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存到内存
    file_buffer = BytesIO()
    df.to_csv(file_buffer, index=False, encoding='utf-8-sig')
    file_buffer.seek(0)
    
    return file_buffer


# 亚马逊模式相关API端点
@product_table_bp.route('/api/parse-amazon-template', methods=['POST'])
def parse_amazon_template():
    """解析亚马逊商品模板文件"""
    try:
        if 'template_file' not in request.files:
            return jsonify({'success': False, 'error': '未找到模板文件'}), 400
        
        file = request.files['template_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'success': False, 'error': '不支持的文件格式，请上传Excel或CSV文件'}), 400
        
        # 解析文件
        variants = parse_template_file(file)
        
        return jsonify({
            'success': True,
            'filename': file.filename,
            'variant_count': len(variants),
            'variants': variants
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@product_table_bp.route('/api/match-variants', methods=['POST'])
def match_variants():
    """匹配变体信息与批次文件夹中的商品"""
    try:
        batch_folder = request.form.get('batch_folder')
        template_file = request.files.get('template_file')
        
        if not batch_folder or not template_file:
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400
        
        # 解析模板文件
        variants = parse_template_file(template_file)
        
        # 获取批次文件夹中的商品
        products = extract_products_from_batch(batch_folder)
        
        # 进行匹配
        matches = match_variants_with_products(variants, products)
        
        matched_count = sum(1 for match in matches if match['product_folder'])
        unmatched_count = len(matches) - matched_count
        
        return jsonify({
            'success': True,
            'matched_count': matched_count,
            'unmatched_count': unmatched_count,
            'matches': matches
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def parse_template_file(file):
    """解析模板文件，提取变体信息"""
    variants = []
    
    try:
        if file.filename.lower().endswith('.csv'):
            # 处理CSV文件
            df = pd.read_csv(file)
        else:
            # 处理Excel文件
            df = pd.read_excel(file)
        
        # 查找关键列（根据常见的亚马逊模板格式）
        key_columns = {
            'parent_sku': None,
            'child_sku': None,
            'color': None,
            'size': None,
            'product_name': None
        }
        
        # 尝试匹配列名（不区分大小写）
        for col in df.columns:
            col_lower = col.lower()
            if 'parent' in col_lower and 'sku' in col_lower:
                key_columns['parent_sku'] = col
            elif 'child' in col_lower and 'sku' in col_lower:
                key_columns['child_sku'] = col
            elif 'color' in col_lower or '颜色' in col_lower:
                key_columns['color'] = col
            elif 'size' in col_lower or '尺寸' in col_lower or '尺码' in col_lower:
                key_columns['size'] = col
            elif 'product' in col_lower and 'name' in col_lower or '商品名称' in col_lower:
                key_columns['product_name'] = col
        
        # 提取变体信息
        for _, row in df.iterrows():
            variant = {}
            for key, col_name in key_columns.items():
                if col_name and col_name in df.columns:
                    variant[key] = str(row[col_name]) if pd.notna(row[col_name]) else ''
            
            if variant:  # 只添加非空的变体
                variants.append(variant)
    
    except Exception as e:
        raise Exception(f"解析模板文件失败: {str(e)}")
    
    return variants


def match_variants_with_products(variants, products):
    """匹配变体信息与商品文件夹"""
    matches = []
    
    # 创建商品名称到文件夹的映射
    product_folders = {product['name'].lower(): product['folder_name'] for product in products}
    
    for variant in variants:
        match = {
            'variant_info': f"{variant.get('parent_sku', '')} - {variant.get('color', '')} - {variant.get('size', '')}",
            'product_folder': None
        }
        
        # 尝试多种匹配策略
        search_terms = []
        
        # 1. 使用商品名称
        if variant.get('product_name'):
            search_terms.append(variant['product_name'].lower())
        
        # 2. 使用颜色信息
        if variant.get('color'):
            search_terms.append(variant['color'].lower())
        
        # 3. 使用SKU信息
        if variant.get('child_sku'):
            search_terms.append(variant['child_sku'].lower())
        if variant.get('parent_sku'):
            search_terms.append(variant['parent_sku'].lower())
        
        # 尝试匹配
        for term in search_terms:
            for product_name, folder_name in product_folders.items():
                if term in product_name or product_name in term:
                    match['product_folder'] = folder_name
                    break
            if match['product_folder']:
                break
        
        matches.append(match)
    
    return matches



@product_table_bp.route('/api/batch-folders')
def get_batch_folders_api():
    """获取批次文件夹列表"""
    try:
        batch_folders = get_batch_folders()
        folders_info = []
        
        for folder in batch_folders:
            folders_info.append({
                'name': folder['name'],
                'product_count': folder['product_count'],
                'created_time': folder['created_time']
            })
        
        return jsonify({
            'folders': folders_info
        })
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


@product_table_bp.route('/api/batch-info/<batch_name>')
def get_batch_info_api(batch_name):
    """获取指定批次的详细信息"""
    try:
        outputs_dir = os.path.join(current_app.root_path, 'static', 'images', 'outputs')
        batch_path = os.path.join(outputs_dir, batch_name)
        
        if not os.path.exists(batch_path):
            return jsonify({'error': '批次文件夹不存在'}), 404
        
        products = extract_products_from_batch(batch_name)
        total_images = sum(product.get('image_count', 0) for product in products)
        
        return jsonify({
            'batch_name': batch_name,
            'product_count': len(products),
            'total_images': total_images,
            'products': [{
                'name': product.get('name', ''),
                'image_count': product.get('image_count', 0),
                'folder_name': product.get('folder_name', ''),
                'preview_image': product.get('preview_image', '')
            } for product in products]
        })
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


def get_batch_folders():
    """获取所有批次文件夹"""
    try:
        outputs_dir = os.path.join(current_app.root_path, 'static', 'images', 'outputs')
        if not os.path.exists(outputs_dir):
            return []
        
        batch_folders = []
        for item in os.listdir(outputs_dir):
            item_path = os.path.join(outputs_dir, item)
            if os.path.isdir(item_path) and item.startswith('batch_'):
                # 获取文件夹信息
                folder_info = {
                    'name': item,
                    'path': item_path,
                    'created_time': datetime.fromtimestamp(os.path.getctime(item_path)).strftime('%Y-%m-%d %H:%M:%S'),
                    'product_count': len([d for d in os.listdir(item_path) if os.path.isdir(os.path.join(item_path, d))])
                }
                batch_folders.append(folder_info)
        
        # 按创建时间倒序排列
        batch_folders.sort(key=lambda x: x['created_time'], reverse=True)
        return batch_folders
        
    except Exception as e:
        print(f"获取批次文件夹时出错: {str(e)}")
        return []


def extract_products_from_batch(batch_name):
    """从批次文件夹中提取商品信息"""
    try:
        outputs_dir = os.path.join(current_app.root_path, 'static', 'images', 'outputs')
        batch_path = os.path.join(outputs_dir, batch_name)
        
        if not os.path.exists(batch_path):
            return []
        
        products = []
        product_folders = [d for d in os.listdir(batch_path) if os.path.isdir(os.path.join(batch_path, d))]
        
        for i, product_folder in enumerate(product_folders, 1):
            product_path = os.path.join(batch_path, product_folder)
            
            # 获取该商品文件夹中的图片文件
            image_files = []
            for ext in ['*.png', '*.jpg', '*.jpeg']:
                image_files.extend(glob.glob(os.path.join(product_path, ext)))
            
            # 生成商品信息
            product_info = {
                'name': product_folder,
                'code': f"P{i:03d}",
                'folder_name': product_folder,
                'batch_name': batch_name,
                'image_count': len(image_files),
                'image_files': [os.path.basename(f) for f in image_files],
                'folder_path': os.path.relpath(product_path, current_app.root_path).replace('\\', '/'),
                'preview_image': os.path.basename(image_files[0]) if image_files else '',
                'created_time': datetime.fromtimestamp(os.path.getctime(product_path)).strftime('%Y-%m-%d %H:%M:%S')
            }
            products.append(product_info)
        
        return products
        
    except Exception as e:
        print(f"提取商品信息时出错: {str(e)}")
        return []


def generate_excel_table_from_folders(title, products):
    """从文件夹信息生成Excel格式的商品表格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"
    
    # 设置样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='微软雅黑', size=11)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ['序号', '商品名称', '商品编码', '文件夹名称', '批次名称', '图片数量', '预览图片', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # 数据行
    for row, product in enumerate(products, 4):
        data = [
            row - 3,  # 序号
            product.get('name', ''),
            product.get('code', ''),
            product.get('folder_name', ''),
            product.get('batch_name', ''),
            product.get('image_count', 0),
            product.get('preview_image', ''),
            product.get('created_time', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = content_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    column_widths = [8, 20, 15, 20, 20, 12, 25, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 保存到内存
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


def generate_csv_table_from_folders(title, products):
    """从文件夹信息生成CSV格式的商品表格"""
    data = []
    for i, product in enumerate(products, 1):
        data.append({
            '序号': i,
            '商品名称': product.get('name', ''),
            '商品编码': product.get('code', ''),
            '文件夹名称': product.get('folder_name', ''),
            '批次名称': product.get('batch_name', ''),
            '图片数量': product.get('image_count', 0),
            '预览图片': product.get('preview_image', ''),
            '创建时间': product.get('created_time', '')
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存到内存
    file_buffer = BytesIO()
    df.to_csv(file_buffer, index=False, encoding='utf-8-sig')
    file_buffer.seek(0)
    
    return file_buffer